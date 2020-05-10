import openpyxl
import click
import os

@click.command()
@click.argument("src",nargs=-1)
@click.argument("dest",nargs=1)
@click.option('--capitalize',default=False,is_flag=True,help="If specified, all string data will be capitalized during copy into destination.")
@click.option('--preservestyles',default=False,is_flag=True,help="If specified, even the cell styles will be copied, else only data.")
def cli(capitalize,preservestyles,src,dest):
    '''Supports some string commands from command line'''
    book = openpyxl.Workbook()
    for s in src:
        t=str(s)
        wb = openpyxl.load_workbook(t)
        sheets = wb.get_sheet_names()
        for n in range(sheets.__len__()):
            sheet=wb.get_sheet_by_name(sheets[n])
            book.create_sheet(sheets[n])
            row=sheet.max_row
            col=sheet.max_column
            click.echo(row)
            click.echo(col)
            sheetdest = book.get_sheet_by_name(sheets[n])
            for i in range(1,row+1):
                for j in range(1,col+1):
                    #click.echo(sheet.cell(row=i, column=j).value)
                    if(capitalize):
                        sheetdest.cell(row=i, column=j).style = sheet.cell(row=i, column=j).style
                        sheetdest.cell(row=i, column=j).value=sheet.cell(row=i, column=j).value.upper()
                    else:
                        sheetdest.cell(row=i, column=j).value = sheet.cell(row=i, column=j).value
    book.save(dest)
    click.echo(sheets)
