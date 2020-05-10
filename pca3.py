import openpyxl
import click
import os
import requests
from bs4 import BeautifulSoup

@click.command()
@click.argument("src",nargs=1)
@click.argument("dest",nargs=1)
def cli(src,dest):
    book = openpyxl.Workbook()
    book.create_sheet('sample')
    sheetdest = book.get_sheet_by_name('sample')
    r=requests.get(src)
    html_content=r.text
    soup=BeautifulSoup(html_content,"html.parser")
    row1=0
    sheetdest.column_dimensions['A'].width =100
    for tr in soup.find_all('tr'):
        row1=row1+1
        col=1
        click.echo(tr)
        for th in tr.find_all('th')[1:]:
            sheetdest.cell(row=row1, column=col).value = th.get_text()
            click.echo(th.get_text())
            col+=1
        break
    for tr in soup.find_all('tr')[1:]:
        row1 = row1 + 1
        col = 1
        for td in tr.find_all('td')[1:]:
            sheetdest.cell(row=row1, column=col).value = td.get_text()
            sheetdest.row_dimensions[row1].height = 20
            col+=1
            #click.echo(td.get_text())
    book.save(dest)
    click.echo('finished')

